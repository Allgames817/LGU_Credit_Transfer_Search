import json
import re
import subprocess
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
COURSES_JS = ROOT / "backend" / "src" / "data" / "courses.js"

XLSX_FILES = [
    Path(r"c:/Users/lenovo/OneDrive - CUHK-Shenzhen/桌面/202603summer equivalent.xlsx"),
    Path(r"c:/Users/lenovo/OneDrive - CUHK-Shenzhen/桌面/202603exchang&visiting equivalent.xlsx"),
]


def norm_str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def code_to_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    return norm_str(v)


def extract_from_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    all_rows: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        header_row = None
        for r in range(1, min(ws.max_row, 120) + 1):
            if norm_str(ws.cell(r, 1).value).lower() == "partner institution":
                header_row = r
                break
        if not header_row:
            continue

        headers: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            h = norm_str(ws.cell(header_row, c).value).lower()
            if h:
                headers[h] = c

        def find_col(contains: str) -> int | None:
            contains = contains.lower()
            for k, col in headers.items():
                if contains in k:
                    return col
            return None

        col_inst = find_col("partner institution") or 1
        col_pcode = find_col("course code")
        col_ptitle = find_col("course title")
        col_ccode = None
        col_ctitle = None
        for k, col in headers.items():
            if "cuhk(sz)" in k and "course code" in k:
                col_ccode = col
            if "cuhk(sz)" in k and "course title" in k:
                col_ctitle = col
        if col_ccode is None:
            col_ccode = find_col("cuhk(sz) course code")
        if col_ctitle is None:
            col_ctitle = find_col("cuhk(sz) course title")
        col_remarks = find_col("remarks")

        current_inst = ""
        for r in range(header_row + 1, ws.max_row + 1):
            inst = norm_str(ws.cell(r, col_inst).value)
            if inst:
                current_inst = inst
            inst = current_inst

            pcode = code_to_str(ws.cell(r, col_pcode).value) if col_pcode else ""
            ptitle = norm_str(ws.cell(r, col_ptitle).value) if col_ptitle else ""
            ccode = code_to_str(ws.cell(r, col_ccode).value) if col_ccode else ""
            ctitle = norm_str(ws.cell(r, col_ctitle).value) if col_ctitle else ""
            remarks = norm_str(ws.cell(r, col_remarks).value) if col_remarks else ""

            if not (inst or pcode or ptitle or ccode or ctitle or remarks):
                continue

            # Only keep real mappings
            if not (inst and pcode and ptitle and ccode and ctitle):
                continue

            all_rows.append(
                {
                    "partnerUniversity": inst,
                    "partnerCourseCode": pcode,
                    "partnerCourseName": ptitle,
                    "cuhkszCourseCode": ccode,
                    "cuhkszCourseName": ctitle,
                    "remarks": remarks,
                }
            )

    return all_rows


def load_existing_courses() -> list[dict]:
    node_cmd = [
        "node",
        "-e",
        "const c=require('./backend/src/data/courses'); console.log(JSON.stringify(c));",
    ]
    existing_json = subprocess.check_output(node_cmd, cwd=str(ROOT), text=True)
    return json.loads(existing_json)


def js_str(s: str) -> str:
    return json.dumps(str(s), ensure_ascii=False)


def fmt_course(c: dict) -> str:
    def num(v) -> int | float:
        try:
            return int(v) if float(v).is_integer() else float(v)
        except Exception:
            return 0

    fields = [
        f"id: {int(c['id'])}",
        f"partnerUniversity: {js_str(c.get('partnerUniversity', ''))}",
        f"partnerCourseCode: {js_str(c.get('partnerCourseCode', ''))}",
        f"partnerCourseName: {js_str(c.get('partnerCourseName', ''))}",
        f"partnerCredits: {num(c.get('partnerCredits', 0))}",
        f"cuhkszCourseCode: {js_str(c.get('cuhkszCourseCode', ''))}",
        f"cuhkszCourseName: {js_str(c.get('cuhkszCourseName', ''))}",
        f"cuhkszCredits: {num(c.get('cuhkszCredits', 0))}",
        f"faculty: {js_str(c.get('faculty', ''))}",
        f"status: {js_str(c.get('status', 'pending'))}",
    ]
    return "  { " + ", ".join(fields) + " }"


def main() -> None:
    for p in XLSX_FILES:
        if not p.exists():
            raise SystemExit(f"Excel not found: {p}")

    existing = load_existing_courses()

    sme_rows: list[dict] = []
    for x in XLSX_FILES:
        sme_rows.extend(extract_from_xlsx(x))

    seen: set[tuple[str, str, str]] = set()
    unique_sme: list[dict] = []
    for r in sme_rows:
        key = (
            norm_str(r["partnerUniversity"]).lower(),
            norm_str(r["partnerCourseCode"]).lower(),
            norm_str(r["cuhkszCourseCode"]).lower(),
        )
        if key in seen:
            continue
        seen.add(key)
        unique_sme.append(r)

    existing_keys = {
        (
            norm_str(c.get("partnerUniversity")).lower(),
            norm_str(c.get("partnerCourseCode")).lower(),
            norm_str(c.get("cuhkszCourseCode")).lower(),
        )
        for c in existing
    }

    # remove any previous SME records to avoid duplicates, then append from Excel
    new_courses = [c for c in existing if norm_str(c.get("faculty")) != "SME"]

    added = 0
    for r in unique_sme:
        key = (
            norm_str(r["partnerUniversity"]).lower(),
            norm_str(r["partnerCourseCode"]).lower(),
            norm_str(r["cuhkszCourseCode"]).lower(),
        )
        if key in existing_keys:
            continue
        new_courses.append(
            {
                "partnerUniversity": r["partnerUniversity"],
                "partnerCourseCode": r["partnerCourseCode"],
                "partnerCourseName": r["partnerCourseName"],
                "partnerCredits": 0,
                "cuhkszCourseCode": r["cuhkszCourseCode"],
                "cuhkszCourseName": r["cuhkszCourseName"],
                "cuhkszCredits": 0,
                "faculty": "SME",
                "status": "approved",
            }
        )
        added += 1

    for i, c in enumerate(new_courses, start=1):
        c["id"] = i

    content = (
        "const courses = [\n"
        + ",\n".join(fmt_course(c) for c in new_courses)
        + "\n];\n\nmodule.exports = courses;\n"
    )
    COURSES_JS.write_text(content, encoding="utf-8")

    sme_count = sum(1 for c in new_courses if c.get("faculty") == "SME")
    print("existing_before", len(existing))
    print("sme_rows_raw", len(sme_rows))
    print("sme_rows_unique", len(unique_sme))
    print("added", added)
    print("final", len(new_courses))
    print("sme_total", sme_count)
    print("final_last_id", new_courses[-1]["id"])


if __name__ == "__main__":
    main()

